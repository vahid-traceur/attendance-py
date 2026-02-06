from flask import Flask, render_template, jsonify, send_file, request
import csv
import os
from datetime import datetime
from collections import Counter, defaultdict
import io
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter

app = Flask(__name__)
FILE_NAME = "attendance.csv"


def read_attendance():
    records = []
    if os.path.exists(FILE_NAME):
        with open(FILE_NAME, "r", encoding="utf-8") as f:
            reader = csv.reader(f)
            for row in reader:
                if len(row) == 3:
                    records.append({
                        "name": row[0],
                        "date": row[1],
                        "time": row[2]
                    })
    return records


def autosize(ws):
    for col in ws.columns:
        max_len = 0
        col_letter = get_column_letter(col[0].column)
        for cell in col:
            if cell.value:
                max_len = max(max_len, len(str(cell.value)))
        ws.column_dimensions[col_letter].width = max_len + 4


@app.route("/")
def dashboard():
    return render_template("dashboard.html")


@app.route("/api/attendance")
def api_attendance():
    records = read_attendance()

    name_filter = request.args.get("name")
    start_date = request.args.get("start")
    end_date = request.args.get("end")
    period = request.args.get("period", "daily")

    if name_filter:
        records = [r for r in records if r["name"] == name_filter]
    if start_date:
        records = [r for r in records if r["date"] >= start_date]
    if end_date:
        records = [r for r in records if r["date"] <= end_date]

    if period == "daily":
        grouped = Counter([r["date"] for r in records])
    elif period == "weekly":
        grouped = defaultdict(int)
        for r in records:
            d = datetime.strptime(r["date"], "%Y-%m-%d")
            y, w, _ = d.isocalendar()
            grouped[f"{y}-W{w}"] += 1
    elif period == "monthly":
        grouped = defaultdict(int)
        for r in records:
            grouped[r["date"][:7]] += 1
    else:
        grouped = {}

    return jsonify({
        "records": records,
        "labels": list(grouped.keys()),
        "values": list(grouped.values())
    })


@app.route("/api/names")
def api_names():
    records = read_attendance()
    return jsonify(sorted(set(r["name"] for r in records)))


@app.route("/download/xlsx")
def download_xlsx():
    records = read_attendance()

    wb = Workbook()

    # -------- Sheet 1: Raw Data --------
    ws = wb.active
    ws.title = "Attendance Data"

    headers = ["Name", "Date", "Time"]
    ws.append(headers)

    header_fill = PatternFill("solid", fgColor="4F81BD")
    header_font = Font(color="FFFFFF", bold=True)

    for col in range(1, 4):
        cell = ws.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    for r in records:
        ws.append([r["name"], r["date"], r["time"]])

    ws.auto_filter.ref = f"A1:C{ws.max_row}"
    ws.freeze_panes = "A2"
    autosize(ws)

    # -------- Sheet 2: Daily Summary --------
    ws2 = wb.create_sheet("Daily Summary")
    ws2.append(["Date", "Attendance Count"])

    daily_counts = Counter([r["date"] for r in records])
    for date, count in sorted(daily_counts.items()):
        ws2.append([date, count])

    for col in range(1, 3):
        cell = ws2.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    ws2.auto_filter.ref = f"A1:B{ws2.max_row}"
    ws2.freeze_panes = "A2"
    autosize(ws2)

    # -------- Sheet 3: Monthly Summary --------
    ws3 = wb.create_sheet("Monthly Summary")
    ws3.append(["Month", "Attendance Count"])

    monthly_counts = defaultdict(int)
    for r in records:
        monthly_counts[r["date"][:7]] += 1

    for month, count in sorted(monthly_counts.items()):
        ws3.append([month, count])

    for col in range(1, 3):
        cell = ws3.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")

    ws3.auto_filter.ref = f"A1:B{ws3.max_row}"
    ws3.freeze_panes = "A2"
    autosize(ws3)

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)

    filename = f"attendance_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
    return send_file(buffer,
                     as_attachment=True,
                     download_name=filename,
                     mimetype="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")


# if __name__ == "__main__":
#     app.run(debug=True)

def run_dashboard():
    app.run(debug=False, use_reloader=False)